"""Generate Interunit Transfers documentation: PDF flow + Excel DB tables."""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT_DIR = os.path.dirname(__file__)

# ────────────────────────────────────────────
# PDF GENERATION
# ────────────────────────────────────────────

class PDF(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 14)
        self.cell(0, 8, "Interunit Transfers - System Flow & API Payloads", align="C", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 8)
        self.cell(0, 5, "Candor Retail  |  Standalone FastAPI App  |  Prefix: /interunit & /transfer", align="C", new_x="LMARGIN", new_y="NEXT")
        self.line(10, self.get_y() + 1, 200, self.get_y() + 1)
        self.ln(4)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "I", 7)
        self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")

    def section(self, title):
        self.set_font("Helvetica", "B", 12)
        self.set_fill_color(230, 240, 255)
        self.cell(0, 8, f"  {title}", fill=True, new_x="LMARGIN", new_y="NEXT")
        self.ln(2)

    def subsection(self, title):
        self.set_font("Helvetica", "B", 10)
        self.set_text_color(30, 80, 160)
        self.cell(0, 6, title, new_x="LMARGIN", new_y="NEXT")
        self.set_text_color(0, 0, 0)
        self.ln(1)

    def body_text(self, text):
        self.set_font("Helvetica", "", 9)
        self.multi_cell(0, 4.5, text)
        self.ln(1)

    def code_block(self, text):
        self.set_font("Courier", "", 7.5)
        self.set_fill_color(245, 245, 245)
        for line in text.split("\n"):
            self.cell(0, 3.8, f"  {line}", fill=True, new_x="LMARGIN", new_y="NEXT")
        self.ln(2)

    def endpoint_row(self, method, path, desc):
        self.set_font("Courier", "B", 8)
        w_method = 18
        w_path = 82
        self.set_fill_color(255, 255, 255)
        color = {"GET": (34,139,34), "POST": (0,100,200), "PATCH": (200,130,0), "PUT": (130,0,180), "DELETE": (200,30,30)}
        r, g, b = color.get(method, (0,0,0))
        self.set_text_color(r, g, b)
        self.cell(w_method, 4.5, method)
        self.set_text_color(0, 0, 0)
        self.set_font("Courier", "", 7.5)
        self.cell(w_path, 4.5, path)
        self.set_font("Helvetica", "", 7.5)
        self.cell(0, 4.5, desc, new_x="LMARGIN", new_y="NEXT")

    def arrow_step(self, step_num, text):
        self.set_font("Helvetica", "B", 9)
        self.cell(8, 5, f"{step_num}.")
        self.set_font("Helvetica", "", 9)
        self.multi_cell(0, 5, text)
        self.ln(0.5)


def build_pdf():
    pdf = PDF()
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # ── 1. Architecture Overview ──
    pdf.section("1. Architecture Overview")
    pdf.body_text(
        "The Interunit Transfers module is a STANDALONE FastAPI application located at "
        "Interunit_transfers/app/main.py. It is NOT mounted inside the main Backend app "
        "(main.py). It connects to the same PostgreSQL (AWS RDS) database but uses its own "
        "set of tables prefixed with 'interunit_transfer_*'.\n\n"
        "Two overlapping router implementations exist:\n"
        "  - /interunit router (interunit.py) - Newer, raw SQL, primary implementation\n"
        "  - /transfer router (transfer.py) - Older, ORM-based, legacy implementation\n\n"
        "Status Flow for Requests:  Pending -> Transferred / Rejected\n"
        "Status Flow for Transfers: Pending -> Partial -> Completed -> Received"
    )

    # ── 2. End-to-End Flow ──
    pdf.section("2. End-to-End Flow (Primary /interunit router)")

    pdf.subsection("Phase A: Transfer Request")
    pdf.arrow_step(1, "User fills form: from_warehouse, to_warehouse, request_date, reason, and adds article lines (material_type, item_category, sub_category, item_description, quantity, uom, pack_size, package_size).")
    pdf.arrow_step(2, "Frontend POSTs to /interunit/requests. Backend generates REQ{YYYYMMDDHHmm} request_no, inserts header into interunit_transfer_requests and lines into interunit_transfer_request_lines. Net weight auto-calculated: pack_size * packaging_type * quantity.")
    pdf.arrow_step(3, "Request appears in list (GET /interunit/requests) with status='Pending'.")
    pdf.arrow_step(4, "Approver can Accept (PATCH status='Accepted') or Reject (PATCH status='Rejected' + reject_reason).")

    pdf.subsection("Phase B: Transfer OUT (Challan / Dispatch)")
    pdf.arrow_step(5, "User creates transfer from an accepted request. POSTs to /interunit/transfers with header (challan_no, stock_trf_date, vehicle_no, driver_name, approved_by, remark, reason_code), lines (same fields as request lines), and optionally scanned boxes.")
    pdf.arrow_step(6, "Backend generates TRANS{YYYYMMDDHHMMSS} challan_no, inserts header into interunit_transfers_header, lines into interunit_transfers_lines, boxes into interunit_transfer_boxes.")
    pdf.arrow_step(7, "If boxes are scanned: compares expected qty (sum of line quantities) vs actual scanned boxes. Sets status='Partial' if boxes pending, 'Completed' if all scanned.")
    pdf.arrow_step(8, "If transfer was created from a request, the request status is updated to 'Transferred'.")

    pdf.subsection("Phase C: Transfer IN (GRN / Receipt)")
    pdf.arrow_step(9, "Receiving warehouse creates Transfer IN by POSTing to /interunit/transfer-in with transfer_out_id, grn_number, receiving_warehouse, received_by, box_condition, and array of scanned_boxes.")
    pdf.arrow_step(10, "Backend inserts header into interunit_transfer_in_header and boxes into interunit_transfer_in_boxes. Updates Transfer OUT status to 'Received'.")
    pdf.arrow_step(11, "Transfer IN can be listed (GET /interunit/transfer-in) and viewed in detail (GET /interunit/transfer-in/{id}).")

    # ── 3. All Endpoints ──
    pdf.add_page()
    pdf.section("3. API Endpoints (/interunit router)")

    pdf.subsection("Dropdown Endpoints")
    pdf.endpoint_row("GET", "/interunit/dropdowns/warehouse-sites", "List warehouse sites")
    pdf.endpoint_row("GET", "/interunit/dropdowns/material-types", "List material types")
    pdf.endpoint_row("GET", "/interunit/dropdowns/units-of-measurement", "List UOMs")
    pdf.endpoint_row("GET", "/interunit/dropdowns/approval-authorities", "List approval authorities")
    pdf.endpoint_row("GET", "/interunit/dropdowns/approval-authorities/warehouse/{wh}", "Authorities by warehouse")
    pdf.ln(2)

    pdf.subsection("Request Endpoints")
    pdf.endpoint_row("POST", "/interunit/requests", "Create transfer request")
    pdf.endpoint_row("GET", "/interunit/requests", "List requests (filters: status, from/to warehouse, created_by)")
    pdf.endpoint_row("GET", "/interunit/requests/{request_id}", "Get request detail with lines")
    pdf.endpoint_row("PATCH", "/interunit/requests/{request_id}", "Update request status (Accept/Reject)")
    pdf.endpoint_row("DELETE", "/interunit/requests/{request_id}", "Delete request (cascade deletes lines)")
    pdf.ln(2)

    pdf.subsection("Transfer OUT Endpoints")
    pdf.endpoint_row("POST", "/interunit/transfers", "Create transfer with lines + boxes")
    pdf.endpoint_row("GET", "/interunit/transfers", "List transfers (paginated, filters: status, site, date, challan)")
    pdf.endpoint_row("GET", "/interunit/transfers/{transfer_id}", "Get transfer detail (header + lines + boxes)")
    pdf.endpoint_row("DELETE", "/interunit/transfers/{transfer_id}", "Delete transfer (only if not Received/Completed)")
    pdf.endpoint_row("PUT", "/interunit/transfers/{transfer_id}/confirm", "Confirm receipt -> status='Received'")
    pdf.ln(2)

    pdf.subsection("Transfer IN (GRN) Endpoints")
    pdf.endpoint_row("POST", "/interunit/transfer-in", "Create Transfer IN with scanned boxes")
    pdf.endpoint_row("GET", "/interunit/transfer-in", "List Transfer INs")
    pdf.endpoint_row("GET", "/interunit/transfer-in/{transfer_in_id}", "Get Transfer IN detail")
    pdf.ln(3)

    # ── 4. Payloads ──
    pdf.add_page()
    pdf.section("4. Request Payloads")

    pdf.subsection("POST /interunit/requests - Create Request")
    pdf.code_block("""{
  "form_data": {
    "request_date": "18-02-2026",          // DD-MM-YYYY
    "from_warehouse": "W202",              // Literal: W202|A185|A101|A68|F53|Savla|Rishi
    "to_warehouse": "A185",
    "reason_description": "MONTHLY STOCK TRANSFER"
  },
  "article_data": [
    {
      "material_type": "RM",               // Literal: RM|PM|FG|RTV
      "item_category": "CHEMICALS",
      "sub_category": "SOLVENTS",
      "item_description": "ACETONE 99%",
      "quantity": "10",
      "uom": "KG",                         // Literal: KG|PCS|BOX|CARTON
      "pack_size": "25.00",
      "package_size": "0",                  // Required only for FG
      "batch_number": "B2026-001",
      "lot_number": "LOT-2026-FEB"
    }
  ],
  "computed_fields": {
    "request_no": "REQ202602181430"         // Optional, auto-generated if null
  },
  "validation_rules": null                  // Optional, frontend-only
}""")

    pdf.subsection("PATCH /interunit/requests/{id} - Update Status")
    pdf.code_block("""{
  "status": "Accepted",                     // or "Rejected"
  "reject_reason": "INSUFFICIENT STOCK",    // Required if rejecting
  "rejected_ts": "2026-02-18T14:30:00"      // Optional
}""")

    pdf.subsection("POST /interunit/transfers - Create Transfer OUT")
    pdf.code_block("""{
  "header": {
    "challan_no": null,                     // Auto-generated TRANS{YYYYMMDDHHMMSS}
    "stock_trf_date": "2026-02-18",         // YYYY-MM-DD (date type)
    "from_warehouse": "W202",
    "to_warehouse": "A185",
    "vehicle_no": "MH12AB1234",
    "driver_name": "RAMESH KUMAR",
    "approved_by": "MANAGER A",
    "remark": "DISPATCHING AS PER REQUEST",
    "reason_code": "MONTHLY TRANSFER"
  },
  "lines": [
    {
      "material_type": "RM",
      "item_category": "CHEMICALS",
      "sub_category": "SOLVENTS",
      "item_description": "ACETONE 99%",
      "quantity": "10",
      "uom": "KG",
      "pack_size": "25.00",
      "package_size": "0",
      "batch_number": "B2026-001",
      "lot_number": "LOT-2026-FEB"
    }
  ],
  "boxes": [
    {
      "box_number": 1,
      "article": "ACETONE 99%",
      "lot_number": "LOT-2026-FEB",
      "batch_number": "B2026-001",
      "transaction_no": "TR-202602181430",
      "net_weight": 25.0,
      "gross_weight": 26.5
    }
  ],
  "request_id": 42                          // Links to original request
}""")

    pdf.add_page()
    pdf.subsection("POST /interunit/transfer-in - Create Transfer IN (GRN)")
    pdf.code_block("""{
  "transfer_out_id": 15,                    // ID of the Transfer OUT header
  "grn_number": "GRN-2026-0218-001",
  "receiving_warehouse": "A185",
  "received_by": "WAREHOUSE SUPERVISOR",
  "box_condition": "Good",                  // Good | Damaged | Partial
  "condition_remarks": null,
  "scanned_boxes": [
    {
      "box_number": "1",
      "article": "ACETONE 99%",
      "batch_number": "B2026-001",
      "lot_number": "LOT-2026-FEB",
      "transaction_no": "TR-202602181430",
      "net_weight": 25.0,
      "gross_weight": 26.5,
      "is_matched": true
    }
  ]
}""")

    # ── 5. Response examples ──
    pdf.section("5. Key Response Structures")

    pdf.subsection("Request Response (RequestWithLines)")
    pdf.code_block("""{
  "id": 1, "request_no": "REQ202602181430",
  "request_date": "18-02-2026",
  "from_warehouse": "W202", "to_warehouse": "A185",
  "reason_description": "MONTHLY STOCK TRANSFER",
  "status": "Pending",
  "reject_reason": null, "created_by": "user@example.com",
  "created_ts": "2026-02-18T14:30:00", "rejected_ts": null, "updated_at": null,
  "lines": [
    { "id": 1, "request_id": 1, "material_type": "RM",
      "item_category": "CHEMICALS", "sub_category": "SOLVENTS",
      "item_description": "ACETONE 99%", "quantity": "10", "uom": "KG",
      "pack_size": "25.00", "package_size": "1", "net_weight": "250.00",
      "batch_number": "B2026-001", "lot_number": "LOT-2026-FEB",
      "created_at": "...", "updated_at": "..." }
  ]
}""")

    pdf.subsection("Transfer List Item (GET /interunit/transfers)")
    pdf.code_block("""{
  "id": 15, "challan_no": "TRANS20260218143055",
  "transfer_no": "TRANS20260218143055",         // alias
  "request_no": "REQ202602181430",
  "stock_trf_date": "18-02-2026",
  "from_site": "W202", "to_site": "A185",
  "vehicle_no": "MH12AB1234", "driver_name": "RAMESH KUMAR",
  "approval_authority": "MANAGER A",
  "status": "Completed",
  "items_count": 1, "boxes_count": 10, "pending_items": 0,
  "has_variance": false
}""")

    # ── 6. Status Flow Diagram ──
    pdf.add_page()
    pdf.section("6. Status Flow Summary")

    pdf.subsection("Transfer Request Status Flow")
    pdf.code_block("""Pending --> Accepted --> Transferred (when Transfer OUT is created)
   |
   +---> Rejected (with reject_reason)""")

    pdf.subsection("Transfer OUT Status Flow")
    pdf.code_block("""Pending --> Partial (some boxes scanned, not all)
   |             |
   |             +---> Completed (all expected boxes scanned)
   |                       |
   +---> Completed -------> Received (Transfer IN / GRN created)""")

    pdf.subsection("Transfer IN Status")
    pdf.code_block("""Created as 'Received' immediately on POST /interunit/transfer-in""")

    pdf.subsection("Net Weight Calculation")
    pdf.body_text(
        "For RM/PM/RTV: net_weight = pack_size * packaging_type * quantity\n"
        "For FG: net_weight = (package_size * pack_size) * quantity\n"
        "Total weight = net_weight * 1.1 (10% packaging overhead) for requests; "
        "same as net_weight for transfers."
    )

    pdf.subsection("Business Rules")
    pdf.body_text(
        "- from_warehouse and to_warehouse must be different.\n"
        "- package_size is required only when material_type is 'FG'.\n"
        "- All text fields (reason, article names, warehouse codes) are uppercased.\n"
        "- Dates: requests use DD-MM-YYYY format, transfers use YYYY-MM-DD (date type).\n"
        "- Transfers with status 'Received' or 'Completed' cannot be deleted.\n"
        "- GRN number must be unique across all Transfer INs.\n"
        "- When a Transfer IN is created, the parent Transfer OUT status -> 'Received'."
    )

    out_path = os.path.join(OUT_DIR, "Interunit_Transfers_Flow.pdf")
    pdf.output(out_path)
    print(f"PDF saved: {out_path}")


# ────────────────────────────────────────────
# EXCEL GENERATION
# ────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

TABLES = [
    {
        "table": "interunit_transfer_requests",
        "description": "Transfer request headers",
        "router": "/interunit",
        "columns": [
            ("id", "SERIAL PK", "Auto-increment primary key"),
            ("request_no", "VARCHAR", "Unique request number (REQ{YYYYMMDDHHMM})"),
            ("request_date", "DATE", "Requested transfer date"),
            ("from_site", "VARCHAR", "Source warehouse code"),
            ("to_site", "VARCHAR", "Destination warehouse code"),
            ("reason_code", "VARCHAR", "Reason / description for transfer"),
            ("remarks", "TEXT", "Additional remarks"),
            ("status", "VARCHAR", "Pending | Accepted | Rejected | Transferred"),
            ("reject_reason", "VARCHAR", "Reason for rejection (nullable)"),
            ("created_by", "VARCHAR", "Email of creator"),
            ("created_ts", "TIMESTAMP", "Creation timestamp"),
            ("rejected_ts", "TIMESTAMP", "Rejection timestamp (nullable)"),
            ("updated_at", "TIMESTAMP", "Last update timestamp"),
        ],
    },
    {
        "table": "interunit_transfer_request_lines",
        "description": "Line items for transfer requests",
        "router": "/interunit",
        "columns": [
            ("id", "SERIAL PK", "Auto-increment primary key"),
            ("request_id", "INTEGER FK", "References interunit_transfer_requests.id (CASCADE)"),
            ("rm_pm_fg_type", "VARCHAR", "Material type: RM | PM | FG | RTV"),
            ("item_category", "VARCHAR", "Item category"),
            ("sub_category", "VARCHAR", "Sub category"),
            ("item_desc_raw", "VARCHAR", "Item description text"),
            ("pack_size", "NUMERIC", "Pack size (kg for RM/PM, gm for FG)"),
            ("qty", "INTEGER", "Quantity of units"),
            ("uom", "VARCHAR", "Unit of measurement: KG | PCS | BOX | CARTON"),
            ("packaging_type", "NUMERIC", "Package size multiplier (for FG)"),
            ("net_weight", "NUMERIC", "Calculated: pack_size * packaging_type * qty"),
            ("total_weight", "NUMERIC", "net_weight * 1.1 (10% overhead)"),
            ("batch_number", "VARCHAR", "Batch number (nullable)"),
            ("lot_number", "VARCHAR", "Lot number (nullable)"),
            ("created_at", "TIMESTAMP", "Row creation timestamp"),
            ("updated_at", "TIMESTAMP", "Row update timestamp"),
        ],
    },
    {
        "table": "interunit_transfers_header",
        "description": "Transfer OUT (challan/dispatch) headers",
        "router": "/interunit",
        "columns": [
            ("id", "SERIAL PK", "Auto-increment primary key"),
            ("challan_no", "VARCHAR UNIQUE", "Challan number (TRANS{YYYYMMDDHHMMSS})"),
            ("stock_trf_date", "DATE", "Transfer date"),
            ("from_site", "VARCHAR", "Source warehouse code"),
            ("to_site", "VARCHAR", "Destination warehouse code"),
            ("vehicle_no", "VARCHAR", "Vehicle registration number"),
            ("driver_name", "VARCHAR", "Driver name"),
            ("approved_by", "VARCHAR", "Approval authority name"),
            ("remark", "TEXT", "Remark / notes"),
            ("reason_code", "VARCHAR", "Reason code for transfer"),
            ("status", "VARCHAR(20)", "Pending | Partial | Completed | Received"),
            ("request_id", "INTEGER FK", "References interunit_transfer_requests.id (nullable)"),
            ("created_by", "VARCHAR", "Email of creator"),
            ("created_ts", "TIMESTAMP", "Creation timestamp"),
            ("approved_ts", "TIMESTAMP", "Approval timestamp (nullable)"),
            ("updated_ts", "TIMESTAMP", "Last update timestamp"),
            ("has_variance", "BOOLEAN", "Whether box count differs from expected"),
        ],
    },
    {
        "table": "interunit_transfers_lines",
        "description": "Line items for Transfer OUT",
        "router": "/interunit",
        "columns": [
            ("id", "SERIAL PK", "Auto-increment primary key"),
            ("header_id", "INTEGER FK", "References interunit_transfers_header.id"),
            ("rm_pm_fg_type", "VARCHAR", "Material type: RM | PM | FG | RTV"),
            ("item_category", "VARCHAR", "Item category"),
            ("sub_category", "VARCHAR", "Sub category"),
            ("item_desc_raw", "VARCHAR", "Item description text"),
            ("item_id", "INTEGER", "SKU / item ID (nullable)"),
            ("hsn_code", "VARCHAR", "HSN code (nullable)"),
            ("pack_size", "NUMERIC", "Pack size"),
            ("packaging_type", "NUMERIC", "Package size multiplier"),
            ("qty", "INTEGER", "Quantity"),
            ("uom", "VARCHAR", "Unit of measurement"),
            ("net_weight", "NUMERIC", "Calculated net weight"),
            ("total_weight", "NUMERIC", "Total weight"),
            ("batch_number", "VARCHAR", "Batch number (nullable)"),
            ("lot_number", "VARCHAR", "Lot number (nullable)"),
            ("created_at", "TIMESTAMP", "Row creation timestamp"),
            ("updated_at", "TIMESTAMP", "Row update timestamp"),
        ],
    },
    {
        "table": "interunit_transfer_boxes",
        "description": "Scanned boxes for Transfer OUT",
        "router": "/interunit",
        "columns": [
            ("id", "SERIAL PK", "Auto-increment primary key"),
            ("header_id", "INTEGER FK", "References interunit_transfers_header.id"),
            ("transfer_line_id", "INTEGER FK", "References interunit_transfers_lines.id"),
            ("box_number", "INTEGER", "Box number (sequential)"),
            ("article", "VARCHAR", "Article / item description"),
            ("lot_number", "VARCHAR", "Lot number (nullable)"),
            ("batch_number", "VARCHAR", "Batch number (nullable)"),
            ("transaction_no", "VARCHAR", "Source transaction number"),
            ("net_weight", "NUMERIC", "Net weight of box"),
            ("gross_weight", "NUMERIC", "Gross weight of box"),
            ("created_at", "TIMESTAMP", "Row creation timestamp"),
            ("updated_at", "TIMESTAMP", "Row update timestamp"),
        ],
    },
    {
        "table": "interunit_transfer_in_header",
        "description": "Transfer IN (GRN / receipt) headers",
        "router": "/interunit",
        "columns": [
            ("id", "SERIAL PK", "Auto-increment primary key"),
            ("transfer_out_id", "INTEGER FK", "References interunit_transfers_header.id"),
            ("transfer_out_no", "VARCHAR", "Challan number of the Transfer OUT"),
            ("grn_number", "VARCHAR UNIQUE", "GRN number (must be unique)"),
            ("grn_date", "TIMESTAMP", "GRN creation date"),
            ("receiving_warehouse", "VARCHAR", "Receiving warehouse code"),
            ("received_by", "VARCHAR", "Person who received"),
            ("received_at", "TIMESTAMP", "Receipt timestamp"),
            ("box_condition", "VARCHAR", "Good | Damaged | Partial"),
            ("condition_remarks", "TEXT", "Remarks about box condition"),
            ("status", "VARCHAR", "Always 'Received' on creation"),
            ("created_at", "TIMESTAMP", "Row creation timestamp"),
            ("updated_at", "TIMESTAMP", "Row update timestamp"),
        ],
    },
    {
        "table": "interunit_transfer_in_boxes",
        "description": "Scanned boxes for Transfer IN (GRN)",
        "router": "/interunit",
        "columns": [
            ("id", "SERIAL PK", "Auto-increment primary key"),
            ("header_id", "INTEGER FK", "References interunit_transfer_in_header.id"),
            ("box_number", "VARCHAR", "Box number"),
            ("article", "VARCHAR", "Article description"),
            ("batch_number", "VARCHAR", "Batch number (nullable)"),
            ("lot_number", "VARCHAR", "Lot number (nullable)"),
            ("transaction_no", "VARCHAR", "Source transaction number"),
            ("net_weight", "NUMERIC", "Net weight"),
            ("gross_weight", "NUMERIC", "Gross weight"),
            ("scanned_at", "TIMESTAMP", "Scan timestamp"),
            ("is_matched", "BOOLEAN", "Whether box matched Transfer OUT data"),
        ],
    },
    {
        "table": "warehouse_sites",
        "description": "Warehouse sites dropdown",
        "router": "/interunit (dropdown)",
        "columns": [
            ("id", "SERIAL PK", "Auto-increment primary key"),
            ("site_code", "VARCHAR", "Warehouse code (e.g. W202, A185)"),
            ("site_name", "VARCHAR", "Display name"),
            ("is_active", "BOOLEAN", "Active status"),
        ],
    },
    {
        "table": "material_types",
        "description": "Material types dropdown",
        "router": "/interunit (dropdown)",
        "columns": [
            ("id", "SERIAL PK", "Auto-increment primary key"),
            ("type_code", "VARCHAR", "Code (RM, PM, FG, RTV)"),
            ("type_name", "VARCHAR", "Display name"),
            ("description", "TEXT", "Description (nullable)"),
            ("is_active", "BOOLEAN", "Active status"),
        ],
    },
    {
        "table": "units_of_measurement",
        "description": "UOM dropdown",
        "router": "/interunit (dropdown)",
        "columns": [
            ("id", "SERIAL PK", "Auto-increment primary key"),
            ("uom_code", "VARCHAR", "Code (KG, PCS, BOX, CARTON)"),
            ("uom_name", "VARCHAR", "Display name"),
            ("description", "TEXT", "Description (nullable)"),
            ("is_active", "BOOLEAN", "Active status"),
        ],
    },
    {
        "table": "transfers_approval_authorities",
        "description": "Approval authorities dropdown",
        "router": "/interunit (dropdown)",
        "columns": [
            ("id", "SERIAL PK", "Auto-increment primary key"),
            ("authority", "VARCHAR", "Authority / person name"),
            ("contact_number", "VARCHAR", "Phone number (nullable)"),
            ("email", "VARCHAR", "Email address (nullable)"),
            ("warehouse", "VARCHAR", "Associated warehouse code"),
            ("is_active", "BOOLEAN", "Active status"),
        ],
    },
    {
        "table": "warehouse_master",
        "description": "Full warehouse address master (ORM, /transfer router)",
        "router": "/transfer (legacy)",
        "columns": [
            ("id", "SERIAL PK", "Auto-increment primary key"),
            ("warehouse_code", "VARCHAR(50) UNIQUE", "Warehouse code"),
            ("warehouse_name", "VARCHAR(200)", "Warehouse name"),
            ("address", "TEXT", "Full address"),
            ("city", "VARCHAR(100)", "City"),
            ("state", "VARCHAR(100)", "State"),
            ("pincode", "VARCHAR(10)", "PIN code"),
            ("gstin", "VARCHAR(15)", "GSTIN number"),
            ("contact_person", "VARCHAR(100)", "Contact person"),
            ("contact_phone", "VARCHAR(15)", "Contact phone"),
            ("contact_email", "VARCHAR(100)", "Contact email"),
            ("is_active", "BOOLEAN", "Active status (default: true)"),
            ("created_at", "TIMESTAMPTZ", "Row creation"),
            ("updated_at", "TIMESTAMPTZ", "Row update"),
        ],
    },
    {
        "table": "transfer_requests",
        "description": "Transfer requests (ORM, /transfer router)",
        "router": "/transfer (legacy)",
        "columns": [
            ("id", "SERIAL PK", "Auto-increment primary key"),
            ("request_no", "VARCHAR(50) UNIQUE", "Request number (REQYYYYMMDDXXX)"),
            ("transfer_no", "VARCHAR(50) UNIQUE", "Transfer number (TRANSYYYYMMDDXXX)"),
            ("request_date", "DATE", "Request date"),
            ("from_warehouse", "VARCHAR(100) FK", "FK -> warehouse_master.warehouse_code"),
            ("to_warehouse", "VARCHAR(100) FK", "FK -> warehouse_master.warehouse_code"),
            ("reason", "VARCHAR(100)", "Short reason"),
            ("reason_description", "TEXT", "Detailed reason"),
            ("status", "VARCHAR(50)", "Pending|Approved|Rejected|In Transit|Completed"),
            ("created_by", "VARCHAR(100)", "Creator email"),
            ("created_at", "TIMESTAMPTZ", "Row creation"),
            ("updated_at", "TIMESTAMPTZ", "Row update"),
        ],
    },
    {
        "table": "transfer_request_items",
        "description": "Items per transfer request (ORM, /transfer router)",
        "router": "/transfer (legacy)",
        "columns": [
            ("id", "SERIAL PK", "Auto-increment primary key"),
            ("transfer_id", "INTEGER FK", "FK -> transfer_requests.id (CASCADE)"),
            ("line_number", "INTEGER", "Line number (unique per transfer)"),
            ("material_type", "VARCHAR(50)", "RM | PM | FG | SFG"),
            ("item_category", "VARCHAR(100)", "Item category"),
            ("sub_category", "VARCHAR(100)", "Sub category"),
            ("item_description", "VARCHAR(500)", "Item description"),
            ("sku_id", "VARCHAR(100)", "SKU ID"),
            ("quantity", "NUMERIC(15,3)", "Quantity"),
            ("uom", "VARCHAR(20)", "Unit of measurement"),
            ("pack_size", "NUMERIC(10,2)", "Pack size"),
            ("package_size", "VARCHAR(50)", "Package size"),
            ("net_weight", "NUMERIC(10,3)", "Net weight"),
            ("created_at", "TIMESTAMPTZ", "Row creation"),
        ],
    },
    {
        "table": "transfer_scanned_boxes",
        "description": "QR-scanned boxes (ORM, /transfer router)",
        "router": "/transfer (legacy)",
        "columns": [
            ("id", "SERIAL PK", "Auto-increment primary key"),
            ("transfer_id", "INTEGER FK", "FK -> transfer_requests.id (CASCADE)"),
            ("box_id", "INTEGER", "Box ID from inward system"),
            ("transaction_no", "VARCHAR(100)", "Source transaction number"),
            ("sku_id", "VARCHAR(100)", "SKU ID"),
            ("box_number_in_array", "INTEGER", "Box index in array"),
            ("box_number", "INTEGER", "Actual box number"),
            ("item_description", "VARCHAR(500)", "Item description"),
            ("net_weight", "NUMERIC(10,3)", "Net weight"),
            ("gross_weight", "NUMERIC(10,3)", "Gross weight"),
            ("scan_timestamp", "TIMESTAMPTZ", "When the box was scanned"),
            ("qr_data", "JSON", "Raw QR data payload"),
        ],
    },
    {
        "table": "transfer_info",
        "description": "Transport details per transfer (ORM, /transfer router)",
        "router": "/transfer (legacy)",
        "columns": [
            ("id", "SERIAL PK", "Auto-increment primary key"),
            ("transfer_id", "INTEGER FK UNIQUE", "FK -> transfer_requests.id (CASCADE, unique)"),
            ("vehicle_number", "VARCHAR(50)", "Vehicle registration"),
            ("vehicle_number_other", "VARCHAR(50)", "Alt vehicle (nullable)"),
            ("driver_name", "VARCHAR(100)", "Driver name"),
            ("driver_name_other", "VARCHAR(100)", "Alt driver (nullable)"),
            ("driver_phone", "VARCHAR(15)", "Driver phone (nullable)"),
            ("approval_authority", "VARCHAR(100)", "Approval authority name"),
            ("created_at", "TIMESTAMPTZ", "Row creation"),
        ],
    },
]


def build_excel():
    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.column_dimensions["A"].width = 6
    ws_summary.column_dimensions["B"].width = 42
    ws_summary.column_dimensions["C"].width = 45
    ws_summary.column_dimensions["D"].width = 25
    ws_summary.column_dimensions["E"].width = 12

    headers = ["#", "Table Name", "Description", "Router", "Columns"]
    for col, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for i, t in enumerate(TABLES, 1):
        row = i + 1
        fill = ALT_FILL if i % 2 == 0 else None
        vals = [i, t["table"], t["description"], t["router"], len(t["columns"])]
        for col, v in enumerate(vals, 1):
            cell = ws_summary.cell(row=row, column=col, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    # ── Sheet per table ──
    for t in TABLES:
        safe_name = t["table"][:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=safe_name)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 55

        # Title row
        ws.merge_cells("A1:C1")
        title_cell = ws.cell(row=1, column=1, value=f"{t['table']}  —  {t['description']}")
        title_cell.font = Font(name="Calibri", bold=True, size=12, color="2F5496")
        title_cell.alignment = Alignment(horizontal="left")

        # Router info
        ws.cell(row=2, column=1, value=f"Router: {t['router']}").font = Font(italic=True, size=9, color="666666")

        # Column headers
        col_headers = ["Column Name", "Data Type", "Description"]
        for col, h in enumerate(col_headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        for i, (col_name, dtype, desc) in enumerate(t["columns"], 1):
            row = i + 4
            fill = ALT_FILL if i % 2 == 0 else None
            for col_idx, val in enumerate([col_name, dtype, desc], 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if fill:
                    cell.fill = fill

    out_path = os.path.join(OUT_DIR, "Interunit_Transfers_DB_Tables.xlsx")
    wb.save(out_path)
    print(f"Excel saved: {out_path}")


if __name__ == "__main__":
    build_pdf()
    build_excel()
    print("Done!")
